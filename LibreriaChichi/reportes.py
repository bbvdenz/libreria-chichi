"""
Generación y envío de reportes de ventas de Librería Chichi.

Módulo reutilizable: lo usan la página de Ventas (descargar/enviar),
el endpoint para el programador automático y el script de respaldo
`enviar_reporte_semanal.py`.

Una "venta" cuenta cuando el pedido está confirmado o entregado.
"""
import io
from datetime import timedelta

from django.utils import timezone
from django.conf import settings
from django.core.mail import EmailMultiAlternatives

ESTADOS_VENTA = ['confirmado', 'entregado']


def clp(v):
    return f"{int(v):,}".replace(",", ".")


def rango_semana(fecha):
    """Devuelve (lunes, domingo) de la semana que contiene 'fecha'."""
    lunes = fecha - timedelta(days=fecha.weekday())
    return lunes, lunes + timedelta(days=6)


def _agrupar(detalles):
    """Devuelve (resumen, por_producto, por_dia, por_semana)."""
    pedidos = set(); unidades = 0; ingresos = 0
    prods = {}; por_dia = {}; por_sem = {}
    for d in detalles:
        sub = int(d.precio_unitario) * d.cantidad
        pedidos.add(d.id_pedido_id); unidades += d.cantidad; ingresos += sub
        k = d.id_producto_id
        prods.setdefault(k, {'nombre': d.id_producto.nombre_producto,
                             'sku': d.id_producto.sku or '-', 'unidades': 0, 'ingresos': 0})
        prods[k]['unidades'] += d.cantidad; prods[k]['ingresos'] += sub
        f = d.id_pedido.fecha_pedido
        por_dia.setdefault(f, {'pedidos': set(), 'unidades': 0, 'ingresos': 0})
        por_dia[f]['pedidos'].add(d.id_pedido_id)
        por_dia[f]['unidades'] += d.cantidad; por_dia[f]['ingresos'] += sub
        iso = f.isocalendar(); ck = (iso[0], iso[1])
        if ck not in por_sem:
            lunes = f - timedelta(days=f.weekday())
            por_sem[ck] = {'inicio': lunes, 'fin': lunes + timedelta(days=6),
                           'pedidos': set(), 'unidades': 0, 'ingresos': 0}
        por_sem[ck]['pedidos'].add(d.id_pedido_id)
        por_sem[ck]['unidades'] += d.cantidad; por_sem[ck]['ingresos'] += sub
    resumen = {'n_ventas': len(pedidos), 'unidades': unidades, 'ingresos': ingresos}
    return resumen, prods, por_dia, por_sem


def _datos(fecha=None):
    """Datos de la semana de 'fecha' (o la actual) + histórico completo."""
    from LibreriaChichi.models import DetallePedido
    if fecha is None:
        fecha = timezone.localdate()
    desde, hasta = rango_semana(fecha)

    det_semana = list(
        DetallePedido.objects.select_related('id_pedido', 'id_producto')
        .filter(id_pedido__estado_pedido__in=ESTADOS_VENTA,
                id_pedido__fecha_pedido__gte=desde,
                id_pedido__fecha_pedido__lte=hasta)
    )
    resumen_sem, prods_sem, _, _ = _agrupar(det_semana)

    det_todos = list(
        DetallePedido.objects.select_related('id_pedido', 'id_producto')
        .filter(id_pedido__estado_pedido__in=ESTADOS_VENTA)
    )
    _, _, por_dia, por_sem = _agrupar(det_todos)

    rango = f"{desde.strftime('%d/%m/%Y')} al {hasta.strftime('%d/%m/%Y')}"
    return {'desde': desde, 'hasta': hasta, 'rango': rango,
            'resumen': resumen_sem, 'prods': prods_sem,
            'por_dia': por_dia, 'por_sem': por_sem}


def construir_excel(prods_semana, por_dia, por_sem, resumen_semana, rango):
    """Construye el Excel (.xlsx) y devuelve los bytes, o None si falta openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    head_font = Font(name='Arial', bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="7C3AED")
    money_fmt = '"$"#,##0'
    titulo_font = Font(name='Arial', bold=True, size=14, color="E8007D")

    def encabezar(ws, headers):
        ws.append(headers)
        for c in ws[ws.max_row]:
            c.font = head_font; c.fill = head_fill
            c.alignment = Alignment(horizontal='center')

    # Hoja 1: Resumen de la semana
    ws = wb.active; ws.title = "Resumen semana"
    ws['A1'] = "Libreria Chichi - Reporte de ventas"; ws['A1'].font = titulo_font
    ws['A2'] = f"Periodo: {rango}"
    ws.append([])
    encabezar(ws, ["Metrica", "Valor"])
    ws.append(["Ventas (pedidos)", resumen_semana['n_ventas']])
    ws.append(["Productos vendidos", resumen_semana['unidades']])
    fila = ws.max_row + 1
    ws.cell(row=fila, column=1, value="Dinero generado")
    ws.cell(row=fila, column=2, value=resumen_semana['ingresos']).number_format = money_fmt
    ws.append([])
    encabezar(ws, ["Producto", "SKU", "Unidades", "Ingresos"])
    for p in sorted(prods_semana.values(), key=lambda x: -x['ingresos']):
        ws.append([p['nombre'], p['sku'], p['unidades'], p['ingresos']])
        ws.cell(row=ws.max_row, column=4).number_format = money_fmt
    ws.column_dimensions['A'].width = 38; ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12; ws.column_dimensions['D'].width = 14

    # Hoja 2: Dia a dia (historico completo)
    ws2 = wb.create_sheet("Dia a dia")
    encabezar(ws2, ["Fecha", "Ventas", "Productos", "Ingresos"])
    for f in sorted(por_dia.keys(), reverse=True):
        v = por_dia[f]
        ws2.append([f.strftime('%d/%m/%Y'), len(v['pedidos']), v['unidades'], v['ingresos']])
        ws2.cell(row=ws2.max_row, column=4).number_format = money_fmt
    for col, w in zip('ABCD', (16, 10, 12, 14)):
        ws2.column_dimensions[col].width = w

    # Hoja 3: Semana a semana (historico completo)
    ws3 = wb.create_sheet("Semana a semana")
    encabezar(ws3, ["Semana (desde - hasta)", "N semana", "Ventas", "Productos", "Ingresos"])
    for ck in sorted(por_sem.keys(), reverse=True):
        v = por_sem[ck]
        rango_s = f"{v['inicio'].strftime('%d/%m/%Y')} - {v['fin'].strftime('%d/%m/%Y')}"
        ws3.append([rango_s, ck[1], len(v['pedidos']), v['unidades'], v['ingresos']])
        ws3.cell(row=ws3.max_row, column=5).number_format = money_fmt
    for col, w in zip('ABCDE', (26, 10, 10, 12, 14)):
        ws3.column_dimensions[col].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def generar_excel(fecha=None):
    """Devuelve (nombre_archivo, bytes|None) del reporte de la semana de 'fecha'."""
    d = _datos(fecha)
    contenido = construir_excel(d['prods'], d['por_dia'], d['por_sem'], d['resumen'], d['rango'])
    nombre = (f"ventas_libreria_chichi_"
              f"{d['desde'].strftime('%Y-%m-%d')}_a_{d['hasta'].strftime('%Y-%m-%d')}.xlsx")
    return nombre, contenido


def _cuerpos_correo(d):
    """Devuelve (texto_plano, html) del correo del reporte."""
    resumen = d['resumen']; rango = d['rango']
    top = sorted(d['prods'].values(), key=lambda x: -x['ingresos'])

    lineas = [
        f"Reporte semanal de ventas - {rango}", "",
        f"Ventas (pedidos): {resumen['n_ventas']}",
        f"Productos vendidos: {resumen['unidades']}",
        f"Dinero generado: ${clp(resumen['ingresos'])}", "",
        "Productos mas vendidos:",
    ]
    lineas += [f"  - {p['unidades']}x {p['nombre']} ({p['sku']}) -> ${clp(p['ingresos'])}"
               for p in top] or ["  (sin ventas)"]
    lineas += ["", "Se adjunta el Excel con el historico completo (dia a dia y semana a semana)."]
    texto = "\n".join(lineas)

    filas = "".join(
        f"<tr><td style='padding:6px 10px;border-bottom:1px solid #eee;'>{p['nombre']}"
        f"<br><small style='color:#888;font-family:monospace;'>{p['sku']}</small></td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eee;text-align:right;'>{p['unidades']}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #eee;text-align:right;color:#10b981;font-weight:700;'>${clp(p['ingresos'])}</td></tr>"
        for p in top
    ) or "<tr><td colspan='3' style='padding:14px;color:#888;text-align:center;'>Sin ventas este periodo.</td></tr>"

    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:auto;color:#222;">
      <h2 style="color:#e8007d;margin-bottom:2px;">Libreria Chichi</h2>
      <p style="color:#666;margin-top:0;">Reporte de ventas - {rango}</p>
      <div style="display:flex;gap:10px;margin:18px 0;">
        <div style="flex:1;background:#faf5ff;border:1px solid #e9d5ff;border-radius:10px;padding:14px;text-align:center;">
          <div style="font-size:0.75rem;color:#7c3aed;font-weight:700;">VENTAS</div>
          <div style="font-size:1.6rem;font-weight:900;">{resumen['n_ventas']}</div></div>
        <div style="flex:1;background:#faf5ff;border:1px solid #e9d5ff;border-radius:10px;padding:14px;text-align:center;">
          <div style="font-size:0.75rem;color:#7c3aed;font-weight:700;">PRODUCTOS</div>
          <div style="font-size:1.6rem;font-weight:900;">{resumen['unidades']}</div></div>
        <div style="flex:1;background:#fdf2f8;border:1px solid #fbcfe8;border-radius:10px;padding:14px;text-align:center;">
          <div style="font-size:0.75rem;color:#e8007d;font-weight:700;">DINERO</div>
          <div style="font-size:1.6rem;font-weight:900;color:#e8007d;">${clp(resumen['ingresos'])}</div></div>
      </div>
      <h3 style="color:#7c3aed;">Productos mas vendidos</h3>
      <table style="width:100%;border-collapse:collapse;font-size:0.9rem;">
        <thead><tr style="background:#7c3aed;color:#fff;">
          <th style="padding:8px 10px;text-align:left;">Producto</th>
          <th style="padding:8px 10px;text-align:right;">Unid.</th>
          <th style="padding:8px 10px;text-align:right;">Ingresos</th></tr></thead>
        <tbody>{filas}</tbody>
      </table>
      <p style="color:#888;font-size:0.85rem;margin-top:16px;">Se adjunta el Excel con el historico completo como respaldo.</p>
    </div>
    """
    return texto, html


def enviar_reporte_correo(destinatario, fecha=None):
    """Genera y envía el reporte de la semana de 'fecha' a 'destinatario'.
    Devuelve True si se envió, False si no (sin lanzar excepción)."""
    if not destinatario:
        print("⚠️ No hay destinatario configurado para el reporte (REPORTE_EMAIL).")
        return False
    d = _datos(fecha)
    texto, html = _cuerpos_correo(d)
    nombre, contenido = generar_excel(fecha)
    remitente = getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@libreriachichi.cl')
    asunto = f"Reporte de ventas Librería Chichi - {d['rango']}"
    try:
        msg = EmailMultiAlternatives(asunto, texto, remitente, [destinatario])
        msg.attach_alternative(html, "text/html")
        if contenido:
            msg.attach(nombre, contenido,
                       'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        msg.send()
        print(f"Reporte enviado a {destinatario} ({d['rango']}).")
        return True
    except Exception as e:
        print(f"⚠️ No se pudo enviar el reporte: {e}")
        return False
